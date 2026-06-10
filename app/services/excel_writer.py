"""
Build the output Excel workbook from processed rows.
Extracted from the original main.py /api/process handler.
"""
import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from app.services.processor import OUTPUT_COLUMNS

INTEGER_COLS = {
    "Advertiser ID", "Insertion Order ID",
    "Impressions", "Billable Impressions", "Clicks",
    "Start Views", "1st Quartile Views", "Midpoint Views",
    "3rd Quartile Views", "Complete Views",
    "Viewable Impressions", "Measurable Impressions",
    "For Checking (Measurable-Impression)", "Start Views-Impression",
}
FLOAT_COLS   = {"Revenue (Adv Currency)", "Media Cost (Advertiser Currency)"}
PERCENT_COLS = {"Click Rate (CTR)", "Video Completion Rate", "Viewability"}


def _is_null(v) -> bool:
    if v is None or v == "":
        return True
    try:
        return math.isnan(float(v)) if isinstance(v, float) else False
    except Exception:
        return False


def build_excel(output_rows: list[dict]) -> io.BytesIO:
    """Convert processed rows into a styled .xlsx BytesIO buffer."""
    out_df = pd.DataFrame(output_rows, columns=OUTPUT_COLUMNS)

    wb = Workbook()
    ws = wb.active

    # Header row
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx in range(len(out_df)):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            raw = out_df.iloc[row_idx][col_name]
            r = row_idx + 2
            if _is_null(raw):
                ws.cell(row=r, column=col_idx, value=None)
            elif col_name in PERCENT_COLS:
                try:
                    num = float(str(raw).replace("%", "").strip()) / 100
                    cell = ws.cell(row=r, column=col_idx, value=num)
                    cell.number_format = "0.00%"
                except Exception:
                    ws.cell(row=r, column=col_idx, value=None)
            elif col_name in INTEGER_COLS:
                try:
                    ws.cell(row=r, column=col_idx, value=int(float(str(raw))))
                except Exception:
                    ws.cell(row=r, column=col_idx, value=None)
            elif col_name in FLOAT_COLS:
                try:
                    ws.cell(row=r, column=col_idx, value=float(str(raw)))
                except Exception:
                    ws.cell(row=r, column=col_idx, value=None)
            else:
                val = str(raw).strip() if raw is not None else None
                ws.cell(row=r, column=col_idx, value=val if val else None)

    # Suppress number-stored-as-text warnings
    try:
        last_col = get_column_letter(len(OUTPUT_COLUMNS))
        last_row = max(len(out_df) + 1, 2)
        ws.ignored_errors.append({"sqref": f"A1:{last_col}{last_row}", "numberStoredAsText": True})
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
