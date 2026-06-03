"""
FastAPI application - CTR processing tool.
"""
import io
import os
import math
from datetime import datetime
from typing import Optional

import pandas as pd
from fastapi import FastAPI, File, UploadFile, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, field_validator
from sqlalchemy.orm import Session

from database import (
    get_db, load_yesterday_memory, save_today_snapshot,
    get_global_settings, get_campaign_rules,
    GlobalSettings, CampaignRule,
)
from processor import process_rows, OUTPUT_COLUMNS

app = FastAPI(title="CTR Processing Tool", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.isdir(FRONTEND_DIR):
    app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

    @app.get("/")
    def serve_frontend():
        return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))


class GlobalSettingsSchema(BaseModel):
    min_ctr: float
    max_ctr: float

    @field_validator('min_ctr', 'max_ctr')
    @classmethod
    def validate_ctr(cls, v):
        if not (0 < v < 100):
            raise ValueError('CTR must be between 0 and 100 (percent)')
        return v


class CampaignRuleSchema(BaseModel):
    line_item_id:  str
    campaign_name: Optional[str] = ''
    min_ctr:       Optional[float] = None
    max_ctr:       Optional[float] = None
    min_vcr:       Optional[float] = None
    max_vcr:       Optional[float] = None
    min_viewability: Optional[float] = None
    max_viewability: Optional[float] = None
    enabled:       Optional[bool] = True

    @field_validator(
        'min_ctr', 'max_ctr',
        'min_vcr', 'max_vcr',
        'min_viewability', 'max_viewability',
    )
    @classmethod
    def validate_ctr(cls, v):
        if v is None:
            return v
        if not (0 < v < 100):
            raise ValueError('Percent values must be between 0 and 100')
        return v


@app.get("/api/settings")
def read_settings(db: Session = Depends(get_db)):
    s = get_global_settings(db)
    return {"min_ctr": s.min_ctr, "max_ctr": s.max_ctr}


@app.put("/api/settings")
def update_settings(payload: GlobalSettingsSchema, db: Session = Depends(get_db)):
    if payload.min_ctr >= payload.max_ctr:
        raise HTTPException(400, "min_ctr must be less than max_ctr")
    s = get_global_settings(db)
    s.min_ctr = payload.min_ctr
    s.max_ctr = payload.max_ctr
    db.commit()
    return {"min_ctr": s.min_ctr, "max_ctr": s.max_ctr}


@app.get("/api/campaigns")
def list_campaigns(db: Session = Depends(get_db)):
    rules = db.query(CampaignRule).order_by(CampaignRule.created_at.desc()).all()
    return [
        {
            "id": r.id,
            "campaign_id": r.line_item_id,
            "campaign_name": r.campaign_name,
            "min_ctr": r.min_ctr,
            "max_ctr": r.max_ctr,
            "min_vcr": r.min_vcr,
            "max_vcr": r.max_vcr,
            "min_viewability": r.min_viewability,
            "max_viewability": r.max_viewability,
            "enabled": r.enabled,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rules
    ]


@app.post("/api/campaigns")
def upsert_campaign(payload: CampaignRuleSchema, db: Session = Depends(get_db)):
    if payload.min_ctr is not None and payload.max_ctr is not None and payload.min_ctr >= payload.max_ctr:
        raise HTTPException(400, "min_ctr must be less than max_ctr")
    if payload.min_vcr is not None and payload.max_vcr is not None and payload.min_vcr > payload.max_vcr:
        raise HTTPException(400, "min_vcr must be less than or equal to max_vcr")
    if payload.min_viewability is not None and payload.max_viewability is not None and payload.min_viewability > payload.max_viewability:
        raise HTTPException(400, "min_viewability must be less than or equal to max_viewability")

    rule = db.query(CampaignRule).filter_by(line_item_id=payload.line_item_id).first()
    if rule:
        rule.campaign_name = payload.campaign_name or ''
        rule.min_ctr = payload.min_ctr if payload.min_ctr is not None else rule.min_ctr
        rule.max_ctr = payload.max_ctr if payload.max_ctr is not None else rule.max_ctr
        rule.min_vcr = payload.min_vcr
        rule.max_vcr = payload.max_vcr
        rule.min_viewability = payload.min_viewability
        rule.max_viewability = payload.max_viewability
        rule.enabled = payload.enabled if payload.enabled is not None else True
        rule.updated_at = datetime.utcnow()
    else:
        rule = CampaignRule(
            line_item_id=payload.line_item_id,
            campaign_name=payload.campaign_name or '',
            min_ctr=payload.min_ctr if payload.min_ctr is not None else 0.37,
            max_ctr=payload.max_ctr if payload.max_ctr is not None else 0.55,
            min_vcr=payload.min_vcr,
            max_vcr=payload.max_vcr,
            min_viewability=payload.min_viewability,
            max_viewability=payload.max_viewability,
            enabled=payload.enabled if payload.enabled is not None else True,
        )
        db.add(rule)
    db.commit()
    db.refresh(rule)
    return {"id": rule.id, "campaign_id": rule.line_item_id,
            "campaign_name": rule.campaign_name, "min_ctr": rule.min_ctr,
            "max_ctr": rule.max_ctr, "min_vcr": rule.min_vcr,
            "max_vcr": rule.max_vcr, "min_viewability": rule.min_viewability,
            "max_viewability": rule.max_viewability, "enabled": rule.enabled}


# by-id routes MUST come before /{campaign_id} to avoid FastAPI route conflict
@app.delete("/api/campaigns/by-id/{rule_id}")
def delete_campaign_by_id(rule_id: int, db: Session = Depends(get_db)):
    rule = db.query(CampaignRule).filter_by(id=rule_id).first()
    if not rule:
        raise HTTPException(404, "Campaign rule not found")
    db.delete(rule)
    db.commit()
    return {"deleted": rule_id}


@app.patch("/api/campaigns/by-id/{rule_id}/toggle")
def toggle_campaign_by_id(rule_id: int, db: Session = Depends(get_db)):
    rule = db.query(CampaignRule).filter_by(id=rule_id).first()
    if not rule:
        raise HTTPException(404, "Campaign rule not found")
    rule.enabled = not rule.enabled
    rule.updated_at = datetime.utcnow()
    db.commit()
    return {"id": rule_id, "enabled": rule.enabled}


@app.post("/api/process")
async def process_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    filename = file.filename or ''
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext not in ('xlsx', 'xls', 'csv'):
        raise HTTPException(400, "Only .xlsx, .xls, or .csv files are supported")

    contents = await file.read()
    try:
        if ext == 'csv':
            df = pd.read_csv(io.BytesIO(contents), dtype=str, keep_default_na=False)
        else:
            df = pd.read_excel(io.BytesIO(contents), dtype=str, keep_default_na=False)
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    df = df.fillna('').replace({'nan': '', 'NaN': '', 'None': '', 'NaT': ''})
    rows = df.to_dict(orient='records')

    yesterday  = load_yesterday_memory(db)
    global_s   = get_global_settings(db)
    camp_rules = get_campaign_rules(db)

    # Campaign names found in file (for UI feedback)
    file_campaign_names = sorted({
        str(r.get('Campaign', '')).strip().lower()
        for r in rows if r.get('Campaign')
    })
    matched_ids = sorted(set(file_campaign_names) & set(camp_rules.keys()))

    output_rows, snapshot = process_rows(
        rows,
        yesterday_memory=yesterday,
        global_min_ctr=global_s.min_ctr,
        global_max_ctr=global_s.max_ctr,
        campaign_ctr_rules=camp_rules,
    )

    save_today_snapshot(db, snapshot)

    out_df = pd.DataFrame(output_rows, columns=OUTPUT_COLUMNS)

    INTEGER_COLS = {
        'Advertiser ID', 'Insertion Order ID',
        'Impressions', 'Billable Impressions', 'Clicks',
        'Start Views', '1st Quartile Views', 'Midpoint Views',
        '3rd Quartile Views', 'Complete Views',
        'Viewable Impressions', 'Measurable Impressions',
        'For Checking (Measurable-Impression)', 'Start Views-Impression',
    }
    FLOAT_COLS   = {'Revenue (Adv Currency)', 'Media Cost (Advertiser Currency)'}
    PERCENT_COLS = {'Click Rate (CTR)', 'Video Completion Rate', 'Viewability'}

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    def is_null(v):
        if v is None or v == '':
            return True
        try:
            return math.isnan(float(v)) if isinstance(v, float) else False
        except Exception:
            return False

    for row_idx in range(len(out_df)):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            raw = out_df.iloc[row_idx][col_name]
            r = row_idx + 2
            if is_null(raw):
                ws.cell(row=r, column=col_idx, value=None)
            elif col_name in PERCENT_COLS:
                try:
                    num = float(str(raw).replace('%', '').strip()) / 100
                    cell = ws.cell(row=r, column=col_idx, value=num)
                    cell.number_format = '0.00%'
                except Exception:
                    ws.cell(row=r, column=col_idx, value=None)
            elif col_name in INTEGER_COLS:
                try:
                    ws.cell(row=r, column=col_idx, value=int(float(str(raw))))
                except Exception:
                    ws.cell(row=r, column=col_idx, value=None)
            elif col_name in FLOAT_COLS:
                try:
                    ws.cell(row=r, column=col_idx, value=float(str(raw)))
                except Exception:
                    ws.cell(row=r, column=col_idx, value=None)
            else:
                val = str(raw).strip() if raw is not None else None
                ws.cell(row=r, column=col_idx, value=val if val else None)

    # Suppress number-stored-as-text warnings if openpyxl supports it
    try:
        from openpyxl.worksheet.cell_range import CellRange
        last_col = get_column_letter(len(OUTPUT_COLUMNS))
        last_row = max(len(out_df) + 1, 2)
        err_range = f"A1:{last_col}{last_row}"
        ws.ignored_errors.append({"sqref": err_range, "numberStoredAsText": True})
    except Exception:
        pass

    output_buf = io.BytesIO()
    wb.save(output_buf)
    output_buf.seek(0)

    out_name = f"processed_{filename.rsplit('.', 1)[0]}.xlsx"
    return StreamingResponse(
        output_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{out_name}"',
            "X-Matched-Rules": ",".join(matched_ids),
            "X-File-Campaign-Names": ",".join(file_campaign_names[:20]),
            "Access-Control-Expose-Headers": "X-Matched-Rules, X-File-Campaign-Names, Content-Disposition",
        },
    )


@app.delete("/api/memory")
def clear_memory(db: Session = Depends(get_db)):
    from database import YesterdayMemory
    db.query(YesterdayMemory).delete()
    db.commit()
    return {"cleared": True}


@app.get("/api/memory/summary")
def memory_summary(db: Session = Depends(get_db)):
    from database import YesterdayMemory
    rows = db.query(YesterdayMemory).all()
    by_id = {}
    for r in rows:
        by_id.setdefault(r.line_item_id, []).append({'clicks': r.clicks, 'ctr': r.ctr})
    return {"line_items": len(by_id), "total_entries": len(rows), "data": by_id}
